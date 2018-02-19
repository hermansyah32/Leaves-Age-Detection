#!/usr/bin/env python
# -*- coding: utf-8 -*-
import sys
sys.path.append('/usr/local/lib/python3.5/site-packages')
import gi
import shutil
import cv2
import os
import copy as cp
import decimal
import openpyxl
from openpyxl import Workbook
from pathlib import Path
import numpy as np
gi.require_version('Gtk', '3.0')
from gi.repository import Gtk, Gdk, GLib, GdkPixbuf
from natsort import natsorted
from PIL import Image
from time import sleep
import smbus
import time



builder = Gtk.Builder()
builder.add_from_file("Interface.glade")

window = builder.get_object("mainMenu")
image = builder.get_object("image1")
lblHasil = builder.get_object("lblHasil")
lblPersamaan = builder.get_object("lblPersamaan")
window.set_position(Gtk.WindowPosition.CENTER)
window.show_all()



class mainWindowHandler:

    global filename
    global foldername
    global tmpfolder
    global inputaction
    global lastRow
    global lastNo

    
    # Define some device parameters
    I2C_ADDR  = 0x3f # I2C device address
    LCD_WIDTH = 16   # Maximum characters per line

    # Define some device constants
    LCD_CHR = 1 # Mode - Sending data
    LCD_CMD = 0 # Mode - Sending command

    LCD_LINE_1 = 0x80 # LCD RAM address for the 1st line
    LCD_LINE_2 = 0xC0 # LCD RAM address for the 2nd line
    LCD_LINE_3 = 0x94 # LCD RAM address for the 3rd line
    LCD_LINE_4 = 0xD4 # LCD RAM address for the 4th line

    LCD_BACKLIGHT  = 0x08  # On
    #LCD_BACKLIGHT = 0x00  # Off

    ENABLE = 0b00000100 # Enable bit

    # Timing constants
    E_PULSE = 0.0005
    E_DELAY = 0.0005

    #Open I2C interface
    #bus = smbus.SMBus(0)  # Rev 1 Pi uses 0
    bus = smbus.SMBus(1) # Rev 2 Pi uses 1

    def __init__(self):
        # Initialise display
        self.lcd_init()

    def lcd_init(self):
      # Initialise display
      self.lcd_byte(0x33,self.LCD_CMD) # 110011 Initialise
      self.lcd_byte(0x32,self.LCD_CMD) # 110010 Initialise
      self.lcd_byte(0x06,self.LCD_CMD) # 000110 Cursor move direction
      self.lcd_byte(0x0C,self.LCD_CMD) # 001100 Display On,Cursor Off, Blink Off 
      self.lcd_byte(0x28,self.LCD_CMD) # 101000 Data length, number of lines, font size
      self.lcd_byte(0x01,self.LCD_CMD) # 000001 Clear display
      time.sleep(self.E_DELAY)

    def lcd_byte(self, bits, mode):
      # Send byte to data pins
      # bits = the data
      # mode = 1 for data
      #        0 for command

      bits_high = mode | (bits & 0xF0) | self.LCD_BACKLIGHT
      bits_low = mode | ((bits<<4) & 0xF0) | self.LCD_BACKLIGHT

      # High bits
      self.bus.write_byte(self.I2C_ADDR, bits_high)
      self.lcd_toggle_enable(bits_high)

      # Low bits
      self.bus.write_byte(self.I2C_ADDR, bits_low)
      self.lcd_toggle_enable(bits_low)

    def lcd_toggle_enable(self, bits):
      # Toggle enable
      time.sleep(self.E_DELAY)
      self.bus.write_byte(self.I2C_ADDR, (bits | self.ENABLE))
      time.sleep(self.E_PULSE)
      self.bus.write_byte(self.I2C_ADDR,(bits & ~self.ENABLE))
      time.sleep(self.E_DELAY)

    def lcd_string(self, message,line):
      # Send string to display

      message = message.ljust(self.LCD_WIDTH," ")

      self.lcd_byte(line, self.LCD_CMD)

      for i in range(self.LCD_WIDTH):
        self.lcd_byte(ord(message[i]),self.LCD_CHR)

    def printkelcd(self, message1, message2):
        self.lcd_string(message1,self.LCD_LINE_1)
        self.lcd_string(message2,self.LCD_LINE_2)
    
    def onDeleteWindow(self, * args):
        Gtk.main_quit(*args)


    def add_filters(self, dialog):
        filter_any = Gtk.FileFilter()
        filter_any.set_name("Any files")
        filter_any.add_pattern("*")
        dialog.add_filter(filter_any)

    def btnInputFile_clicked_cb(self, button):
        self.inputaction = 2
        dialog = Gtk.FileChooserDialog("Please choose your file", window,
                                       Gtk.FileChooserAction.OPEN, (Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL,
                                                                   Gtk.STOCK_OPEN, Gtk.ResponseType.OK))
        self.add_filters(dialog)
        response = dialog.run()
        if (response == Gtk. ResponseType.OK):
            self.filename = dialog.get_filename()
            image.set_from_file(self.filename)
        dialog.destroy()

    def btnInputFolder_clicked_cb(self, button):
        self.inputaction=1
        dialog = Gtk.FileChooserDialog("Please choose your folder", window,
                                       Gtk.FileChooserAction.SELECT_FOLDER, (Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL,
                                                                   Gtk.STOCK_OPEN, Gtk.ResponseType.OK))
        
        response = dialog.run()
        if (response == Gtk. ResponseType.OK):
            self.foldername = dialog.get_filename()
        dialog.destroy()
        
    def btnInputCamera_clicked_cb(self, button):
        self.inputaction=3
        print("Gambar telah diambil")

        
    def savetoXLSX(self):
        wb = Workbook()
        

        excel_data = Path("data/data.xlsx")
        if excel_data.is_file():
            book = openpyxl.load_workbook('data/data.xlsx')

            sheet = book.active
            self.lastRow = sheet['C2'].value
            self.lastNo = sheet['C3'].value
            path=os.getcwd()
            path = path.replace(" ", "%20")
            dirFiles = natsorted(os.listdir(self.foldername))

            if self.inputaction==1:
                for file in dirFiles:
                    if (file.endswith(".png") or file.endswith(".jpg") or file.endswith(".jpeg")):
                        
                        filename = os.path.join(self.foldername+"/"+file)
                        print ("Processing file: " + filename)
                        shutil.copyfile(filename, "data/sumber/"+str(self.lastNo)+".png")
                        luasDaun = self.countAreaObject(filename, True)
                        rgbDaun = self.RGBCount(filename, True)
                        kelilingDaun = self.countKeliling(filename, True)
                        persamaanluas = self.persamaan_luas(luasDaun)
                        persamaankeliling = self.persamaan_keliling(kelilingDaun)
                        persamaanrgb = self.persamaan_rgb(rgbDaun)
                        umurluas = self.cekumur_luas(persamaanluas)
                        umurkeliling = self.cekumur_keliling(persamaankeliling)
                        umurrgb = self.cekumur_rgb(persamaanrgb)

                        sheet['A'+str(self.lastRow)]=self.lastNo
                        sheet['B'+str(self.lastRow)]=filename
                        #Luas
                        sheet['C'+str(self.lastRow)]=luasDaun[0]
                        sheet['D'+str(self.lastRow)]=luasDaun[1]
                        sheet['E'+str(self.lastRow)]="=HYPERLINK(\"file://"+path+"/data/luasdaun/"+str(self.lastNo)+"_atas.png\",\"LINK\")"
                        sheet['F'+str(self.lastRow)]="=HYPERLINK(\"file://"+path+"/data/luasdaun/"+str(self.lastNo)+"_bawah.png\",\"LINK\")"
                        sheet['AP'+str(self.lastRow)] = persamaanluas
                        sheet['AS'+str(self.lastRow)] = umurluas

                        #RGB
                        colRgb = 7
                        for i in rgbDaun:
                            item = i.tolist()
                            for x in item:
                                sheet.cell(row=self.lastRow, column=colRgb).value = x
                                colRgb +=1

                        sheet['AK'+str(self.lastRow)]="=HYPERLINK(\"file://"+path+"/data/rgbdaun/"+str(self.lastNo)+".png\",\"LINK\")"
                        sheet['AR'+str(self.lastRow)] = persamaanrgb
                        sheet['AU'+str(self.lastRow)] = umurrgb

                        #Keliling
                        sheet['AL'+str(self.lastRow)]=kelilingDaun[0]
                        sheet['AM'+str(self.lastRow)]=kelilingDaun[1]
                        sheet['AN'+str(self.lastRow)]="=HYPERLINK(\"file://"+path+"/data/kelilingdaun/"+str(self.lastNo)+"_atas.png\",\"LINK\")"
                        sheet['AO'+str(self.lastRow)]="=HYPERLINK(\"file://"+path+"/data/kelilingdaun/"+str(self.lastNo)+"_bawah.png\",\"LINK\")"
                        sheet['AQ'+str(self.lastRow)] = persamaankeliling
                        sheet['AT'+str(self.lastRow)] = umurkeliling
                        self.lastRow +=1
                        self.lastNo +=1
                        print ("\n")
                sheet['C2'] = self.lastRow
                sheet['C3'] = self.lastNo
                book.save("data/data.xlsx")
            else:
                shutil.copyfile(self.filename, "data/sumber/"+str(self.lastNo)+".png")
                luasDaun = self.countAreaObject(self.filename, True)
                rgbDaun = self.RGBCount(self.filename, True)
                kelilingDaun = self.countKeliling(self.filename, True)
                sheet['A'+str(self.lastRow)]=self.lastNo
                sheet['B'+str(self.lastRow)]=self.filename
                #Luas
                sheet['C'+str(self.lastRow)]=luasDaun[0]
                sheet['D'+str(self.lastRow)]=luasDaun[1]
                sheet['E'+str(self.lastRow)]="=HYPERLINK(\"file://"+path+"/data/luasdaun/"+str(self.lastNo)+"_atas.png\",\"LINK\")"
                sheet['F'+str(self.lastRow)]="=HYPERLINK(\"file://"+path+"/data/luasdaun/"+str(self.lastNo)+"_bawah.png\",\"LINK\")"

                #RGB
                colRgb = 7
                for i in rgbDaun:
                    item = i.tolist()
                    for x in item:
                        sheet.cell(row=self.lastRow, column=colRgb).value = x
                        colRgb +=1

                sheet['AK'+str(self.lastRow)]="=HYPERLINK(\"file://"+path+"/data/rgbdaun/"+str(self.lastNo)+".png\",\"LINK\")"
                #Keliling
                sheet['AL'+str(self.lastRow)]=kelilingDaun[0]
                sheet['AM'+str(self.lastRow)]=kelilingDaun[1]
                sheet['AN'+str(self.lastRow)]="=HYPERLINK(\"file://"+path+"/data/kelilingdaun/"+str(self.lastNo)+"_atas.png\",\"LINK\")"
                sheet['AO'+str(self.lastRow)]="=HYPERLINK(\"file://"+path+"/data/kelilingdaun/"+str(self.lastNo)+"_bawah.png\",\"LINK\")"
                self.lastRow +=1
                self.lastNo +=1
                sheet['C2'] = self.lastRow
                sheet['C3'] = self.lastNo
                book.save("data/data.xlsx")


    def btnSave_clicked_cb(self, button):
        self.savetoXLSX()
        print ("selesai menyimpan")
        
    def on_btnLuasDaun_clicked(self, button):
        if (self.inputaction==1):
            #inputfromfolder
            print("test")
        else:
            jumlahluas = self.countAreaObject(self.filename, False)
            persamaanluas = self.persamaan_luas(jumlahluas)
            umurluas = self.cekumur_luas(persamaanluas)
            lblPersamaan.set_text("(" + str(umurluas) + ")")
        
    def on_btnRgbDaun_clicked(self, button):
        if (self.inputaction==1):
            #inputfromfolder
            print("test")
        else:
            jumlahrgb = self.RGBCount(self.filename, False)
            persamaanrgb = self.persamaan_rgb(jumlahrgb)
            umurrgb = self.cekumur_rgb(persamaanrgb)
            lblPersamaan.set_text("(" + str(umurrgb) + ")")

    def on_btnKelilingDaun_clicked(self, button):
        if (self.inputaction==1):
            #inputfromfolder
            print("test")
        else:
            jumlahkeliling = self.countKeliling(self.filename, False)
            persamaankeliling = self.persamaan_keliling(jumlahkeliling)
            umurkeliling = self.cekumur_keliling(persamaankeliling)
            lblPersamaan.set_text("(" + str(umurkeliling) + ")")

    def cekumur_keliling(self, inputvar):
        hasil = 0
        #umur 1 bulan
        if (inputvar>=0.7 and inputvar<=0.71):
            hasil = 1
        if (inputvar>=0.749000001 and inputvar<=0.77):
            hasil = 1
        if (inputvar>=0.7460000010 and inputvar<=0.8100000000):
            hasil = 1
        if (inputvar>=0.9250000100 and inputvar<=0.9370000000):
            hasil = 1

        #umur 2 bulan
        if (inputvar>=0.7100000100 and inputvar<=0.7140000000):
            hasil = 2
        if (inputvar>=0.7270000010 and inputvar<=0.7490000000):
            hasil = 2
        if (inputvar>=0.9000000100 and inputvar<=0.9250000000):
            hasil = 2
        if (inputvar>=0.9383500010 and inputvar<=0.9390000000):
            hasil = 2
            
        #umur 3 bulan
        if (inputvar>=0.7140000001 and inputvar<=0.7150000000):
            hasil = 3
        if (inputvar>=0.7190000010 and inputvar<=0.7270000000):
            hasil = 3
        if (inputvar>=0.8500000100 and inputvar<=0.9000000000):
            hasil = 3
        if (inputvar>=0.9370000010 and inputvar<=0.9383500000):
            hasil = 3
        if (inputvar>=0.9390000010 and inputvar<=0.9392000000):
            hasil = 3

        #umur 4 bulan
        if (inputvar>=0.7150000010 and inputvar<=0.7155000000):
            hasil = 4
        if (inputvar>=0.7170000010 and inputvar<=0.7190000000):
            hasil = 4
        if (inputvar>=0.8100000100 and inputvar<=0.8500000000):
            hasil = 4
        if (inputvar>=0.9392500010 and inputvar<=0.9400000000):
            hasil = 4
            
        #umur 5 bulan
        if (inputvar>=0.7155000001 and inputvar<=0.7170000000):
            hasil = 5
        if (inputvar>=0.7700000010 and inputvar<=0.7460000000):
            hasil = 5
        if (inputvar>=0.9392000001 and inputvar<=0.9392500000):
            hasil = 5
        
        print ("Umur daun adalah: " + str(hasil))
        self.printkelcd("Umur daun adalah", str(hasil) + " Bulan")
        return hasil

    def cekumur_luas(self, inputvar):
        hasil = 0
        
        #umur 1 bulan
        if (inputvar>=0.998500000001000 and inputvar<=0.998800000000000):
            hasil = 1
        if (inputvar>=0.999000000000000 and inputvar<=1.000000000000000):
            hasil = 1
        
        #umur 2 bulan
        if (inputvar>=0.990355240000000 and inputvar<=0.990355248900000):
            hasil = 2
        if (inputvar>=0.990355255000001 and inputvar<=0.990355256500000):
            hasil = 2
        if (inputvar>=0.990355500000000 and inputvar<=0.990355862000000):
            hasil = 2
        if (inputvar>=0.990356000001000 and inputvar<=0.990356500000000):
            hasil = 2
        if (inputvar>=0.990358000001000 and inputvar<=0.990450000000000):
            hasil = 2
        if (inputvar>=0.990500000001000 and inputvar<=0.996000000000000):
            hasil = 2
        if (inputvar>=0.996500000010000 and inputvar<=0.998500000000000):
            hasil = 2
        if (inputvar>=0.998800000001000 and inputvar<=0.999000000000000):
            hasil = 2

        #umur 3 bulan
        if (inputvar>=0.990355249000000 and inputvar<=0.990355249250000):
            hasil = 3
        if (inputvar>=0.990355249303001 and inputvar<=0.990355249305000):
            hasil = 3
        if (inputvar>=0.990355249430001 and inputvar<=0.990355249450000):
            hasil = 3
        if (inputvar>=0.990355250000000 and inputvar<=0.990355253000000):
            hasil = 3
        if (inputvar>=0.990355256500010 and inputvar<=0.990355259000000):
            hasil = 3
        if (inputvar>=0.990355260000100 and inputvar<=0.990355499999000):
            hasil = 3
        if (inputvar>=0.990355862000001 and inputvar<=0.990356000000000):
            hasil = 3
        if (inputvar>=0.990356500000010 and inputvar<=0.990358000000000):
            hasil = 3
        if (inputvar>=0.990451000000000 and inputvar<=0.990500000000000):
            hasil = 3

        #umur 4 bulan
        if (inputvar>=0.990355249260000 and inputvar<=0.990355249266999):
            hasil = 4
        if (inputvar>=0.990355249267111 and inputvar<=0.990355249267600):
            hasil = 4
        if (inputvar>=0.990355249269000 and inputvar<=0.990355249271000):
            hasil = 4
        if (inputvar>=0.990355249281001 and inputvar<=0.990355249283000):
            hasil = 4
        if (inputvar>=0.990355249285001 and inputvar<=0.990355249303000):
            hasil = 4
        if (inputvar>=0.990355249305001 and inputvar<=0.990355249340000):
            hasil = 4
        if (inputvar>=0.990355249400010 and inputvar<=0.990355249430000):
            hasil = 4
        if (inputvar>=0.990355249480001 and inputvar<=0.990355250000000):
            hasil = 4
        if (inputvar>=0.990355253000010 and inputvar<=0.990355255000000):
            hasil = 4
        if (inputvar>=0.996000000010000 and inputvar<=0.996500000000000):
            hasil = 4

        #umur 5 bulan
        if (inputvar>=0.990355249267000 and inputvar<=0.990355249267110):
            hasil = 5
        if (inputvar>=0.990355249267601 and inputvar<=0.990355249268999):
            hasil = 5
        if (inputvar>=0.990355249271001 and inputvar<=0.990355249281000):
            hasil = 5
        if (inputvar>=0.990355249283001 and inputvar<=0.990355249285000):
            hasil = 5
        if (inputvar>=0.990355249340010 and inputvar<=0.990355249400000):
            hasil = 5
        if (inputvar>=0.990355249450010 and inputvar<=0.990355249480000):
            hasil = 5
        if (inputvar>=0.990355259000010 and inputvar<=0.990355260000000):
            hasil = 5

        print ("Umur daun adalah: " + str(hasil))
        return hasil

    def cekumur_rgb(self, inputvar):
        inputvar = inputvar*100000
        hasil = 0

        #umur 1 bulan
        if (inputvar>=0.029369724484205 and inputvar<=0.029719015774401):
            hasil = 1
        if (inputvar>=0.030505103311023 and inputvar<=0.0325):
            hasil = 1
        if (inputvar>=0.039230750612453 and inputvar<=0.042599):
            hasil = 1
        if (inputvar>=0.145315385263545 and inputvar<=0.211297905573164):
            hasil = 1
        if (inputvar>=0.214007566327266 and inputvar<=0.217214368282262):
            hasil = 1
        if (inputvar>=0.21775070875511 and inputvar<=0.219367890609806):
            hasil = 1

        #umur 2 bulan
        if (inputvar>=0.0326 and inputvar<=0.034066493664224):
            hasil = 2
        if (inputvar>=0.055999 and inputvar<=0.063999):
            hasil = 2
        if (inputvar>=0.079999 and inputvar<=0.139999):
            hasil = 2
        if (inputvar>=0.211586807805377 and inputvar<=0.213999):
            hasil = 2
        if (inputvar>=0.217493983284268 and inputvar<=0.217699):
            hasil = 2
        if (inputvar>=0.219406643722113 and inputvar<=0.219799):
            hasil = 2
        if (inputvar>=801 and inputvar<=1090):
            hasil = 2
        if (inputvar>=1250 and inputvar<=2000):
            hasil = 2
        if (inputvar>=3562.99 and inputvar<=3566):
            hasil = 2
        if (inputvar>=3610 and inputvar<=3610.469):
            hasil = 2
        if (inputvar>=3611 and inputvar<=3627):
            hasil = 2
        if (inputvar>=3656 and inputvar<=3668):
            hasil = 2
        if (inputvar>=3669 and inputvar<=12000):
            hasil = 2

        #umur 3 bulan
        if (inputvar>=0.02 and inputvar<=0.021):
            hasil = 3
        if (inputvar>=0.0273 and inputvar<=0.0285):
            hasil = 3
        if (inputvar>=2 and inputvar<=80):
            hasil = 3
        if (inputvar>=2101 and inputvar<=3220):
            hasil = 3
        if (inputvar>=3261 and inputvar<=3295):
            hasil = 3
        if (inputvar>=3321 and inputvar<=3370):
            hasil = 3
        if (inputvar>=3421 and inputvar<=3563):
            hasil = 3
        if (inputvar>=3583 and inputvar<=3609):
            hasil = 3
        if (inputvar>=3609.5 and inputvar<=3609.9):
            hasil = 3
        if (inputvar>=3610.47 and inputvar<=3611):
            hasil = 3
        if (inputvar>=3627.1 and inputvar<=3639.999):
            hasil = 3
        if (inputvar>=3663 and inputvar<=3668.999):
            hasil = 3
            
        #umur 4 bulan
        if (inputvar>=0.0001 and inputvar<=0.0104):
            hasil = 4
        if (inputvar>=0.0108 and inputvar<=0.012):
            hasil = 4
        if (inputvar>=0.0137 and inputvar<=0.0138):
            hasil = 4
        if (inputvar>=0.021 and inputvar<=0.0215):
            hasil = 4
        if (inputvar>=0.022 and inputvar<=0.0225):
            hasil = 4
        if (inputvar>=0.026 and inputvar<=0.0269):
            hasil = 4
        if (inputvar>=0.0285 and inputvar<=0.02935):
            hasil = 4
        if (inputvar>=0.03 and inputvar<=0.0305):
            hasil = 4
        if (inputvar>=81 and inputvar<=800):
            hasil = 4
        if (inputvar>=1091 and inputvar<=1249):
            hasil = 4
        if (inputvar>=3221 and inputvar<=3260):
            hasil = 4
        if (inputvar>=3371 and inputvar<=3420):
            hasil = 4
        if (inputvar>=3566.1 and inputvar<=3582.99):
            hasil = 4
        if (inputvar>=3609.1 and inputvar<=3609.499):
            hasil = 4
        if (inputvar>=3609.91 and inputvar<=3609.9):
            hasil = 4

        #umur 5 bulan
        if (inputvar>=0.01199 and inputvar<=0.013699):
            hasil = 5
        if (inputvar>=0.01381 and inputvar<=0.019999):
            hasil = 5
        if (inputvar>=0.02151 and inputvar<=0.021999):
            hasil = 5
        if (inputvar>=0.02551 and inputvar<=0.02559):
            hasil = 5
        if (inputvar>=0.027 and inputvar<=0.027299):
            hasil = 5
        if (inputvar>=0.034067 and inputvar<=0.039):
            hasil = 5
        if (inputvar>=0.0426 and inputvar<=0.055998):
            hasil = 5
        if (inputvar>=0.064 and inputvar<=0.07998):
            hasil = 5
        if (inputvar>=0.2198 and inputvar<=1.99):
            hasil = 5
        if (inputvar>=3628 and inputvar<=3655.99):
            hasil = 5
        print ("Umur daun adalah: " + str(hasil))
        return hasil
            
    def persamaan_rgb(self, inputvar):
        #tahap penjumlahan
        inputvar2 = cp.deepcopy(inputvar)
        indexItem = 0
        for i in inputvar:
            item = i.tolist()
            sumofItem = 0
            for x in item:
                x1= x*0.01
                sumofItem = sumofItem + x1
            inputvar2[indexItem]=sumofItem
            indexItem +=1
            
        v11=1.4004 
        v21=-0.2972 
        v31=1.5537 
        v41=-1.2604 
        v51=0.6302
        v12=-3.5916 
        v22=0.7617 
        v32=-2.7840 
        v42=2.2625 
        v52=2.0473
        v13=1.2331 
        v23=1.5439 
        v33=-1.9324 
        v43=7.1158 
        v53=1.9876
        v14=0.8985 
        v24=-0.0452 
        v34=-0.9155 
        v44=-2.7155 
        v54=-1.8111
        v15=1.8386 
        v25=-1.0646 
        v35=-1.0175 
        v45=-0.2514 
        v55=0.3470
        v16=-1.1064 
        v26=4.5201 
        v36=-1.1779 
        v46=3.0497 
        v56=-0.4884
        v17=-3.3348 
        v27=-0.2015 
        v37=2.0500 
        v47=-0.0438 
        v57=-0.3871
        v18=0.7826 
        v28=0.6025 
        v38=-3.0437 
        v48=1.9025 
        v58=1.7143
        v19=1.1794 
        v29=-2.5524 
        v39=-5.2434 
        v49=-0.2866 
        v59=2.0611
        v110=-2.9668 
        v210=-3.6863 
        v310=0.5320 
        v410=-1.7164 
        v510=0.4906
        bz1=0.8734 
        bz2=-1.7877 
        bz3=1.0375 
        bz4=-5.1055 
        bz5=-5.6914
        w1=-3.3221 
        w2=-0.5377 
        w3=0.7797 
        w4=-1.7300 
        w5=6.3210
        b0=4.2648
        e=10
        
        #tahap pembobotan
        z_in1=w1+((inputvar2[0]*v11)+(inputvar2[1]*v12)+(inputvar2[2]*v13)+(inputvar2[3]*v14)+(inputvar2[4]*v15)+(inputvar2[5]*v16)+(inputvar2[6]*v17)+(inputvar2[7]*v18)+(inputvar2[8]*v19)+(inputvar2[9]*v110))
        z_in2=w2+((inputvar2[0]*v21)+(inputvar2[1]*v22)+(inputvar2[2]*v23)+(inputvar2[3]*v24)+(inputvar2[4]*v25)+(inputvar2[5]*v26)+(inputvar2[6]*v27)+(inputvar2[7]*v28)+(inputvar2[8]*v29)+(inputvar2[9]*v210))
        z_in3=w3+((inputvar2[0]*v31)+(inputvar2[1]*v32)+(inputvar2[2]*v33)+(inputvar2[3]*v34)+(inputvar2[4]*v35)+(inputvar2[5]*v36)+(inputvar2[6]*v37)+(inputvar2[7]*v38)+(inputvar2[8]*v39)+(inputvar2[9]*v310))
        z_in4=w4+((inputvar2[0]*v41)+(inputvar2[1]*v42)+(inputvar2[2]*v43)+(inputvar2[3]*v44)+(inputvar2[4]*v45)+(inputvar2[5]*v46)+(inputvar2[6]*v47)+(inputvar2[7]*v48)+(inputvar2[8]*v49)+(inputvar2[9]*v410))
        z_in5=w5+((inputvar2[0]*v51)+(inputvar2[1]*v52)+(inputvar2[2]*v53)+(inputvar2[3]*v54)+(inputvar2[4]*v55)+(inputvar2[5]*v56)+(inputvar2[6]*v57)+(inputvar2[7]*v58)+(inputvar2[8]*v59)+(inputvar2[9]*v510))
        
        #fungsi aktivasi
        z1=1/(1+pow(e,-(z_in1)))
        z2=1/(1+pow(e,-(z_in2)))
        z3=1/(1+pow(e,-(z_in3)))
        z4=1/(1+pow(e,-(z_in4)))
        z5=1/(1+pow(e,-(z_in5)))
        
        
        #output
        y_in=b0+((z1*bz1)+(z2*bz2)+(z3*bz3)+(z4*bz4)+(z5*bz5))
        result=1/(1+pow(e,-(y_in)))
        print("Hasilt persamaan: " + str(result))
        return result

    def persamaan_keliling(self, inputvar):
        var11=2.0157
        var21=-2.5408
        var31=2.9785
        var41=0.0689
        var51=-4.3687
        var12=-1.8502
        var22=0.5244
        var32=0.3451
        var42=0.1044
        var52=2.0059
        
        bz1=-0.7890
        bz2=0.4114
        bz3=-1.6262
        bz4=3.3549
        bz5=-0.3437
        
        w1=-2.4535
        w2=-2.9795
        w3=3.2752
        w4=0.8474
        w5=4.2016
        
        b0=-0.5387
        e=10


        #tahap pembobotan
        z_in1=w1+((inputvar[0]*var11)+(inputvar[1]*var12))
        z_in2=w2+((inputvar[0]*var21)+(inputvar[1]*var22))
        z_in3=w3+((inputvar[0]*var31)+(inputvar[1]*var32))
        z_in4=w4+((inputvar[0]*var41)+(inputvar[1]*var42))
        z_in5=w5+((inputvar[0]*var51)+(inputvar[1]*var52))

        #fungsi aktivasi
        z1=1/(1+pow(e,-(z_in1)))
        z2=1/(1+pow(e,-(z_in2)))
        z3=1/(1+pow(e,-(z_in3)))
        z4=1/(1+pow(e,-(z_in4)))
        z5=1/(1+pow(e,-(z_in5)))

        #output
        y_in=b0+((z1*bz1)+(z2*bz2)+(z3*bz3)+(z4*bz4)+(z5*bz5))
        result=1/(1+pow(e,-(y_in)))
        print("Hasilt persamaan: " + str(result))
        return result
        

    def persamaan_luas(self, inputvar):
        inputvar[0] = inputvar[0]*0.1
        inputvar[1] = inputvar[1]*0.1
        v11=0.5430 
        v21=1.6464 
        v31=-5.5055 
        v41=0.6884 
        v51=-0.5672
        v12=6.3309 
        v22=-4.7068 
        v32=-7.0952 
        v42=-5.5650 
        v52=-1.6284
        bz1=-0.7986 
        bz2=1.6986 
        bz3=-3.3620 
        bz4=1.4270 
        bz5=-9.4362
        w1=-6.6344 
        w2=-1.6962 
        w3=4.1913 
        w4=6.2366 
        w5=-2.4389
        b0=2.8101
        e=10
        
        #tahap pembobotan
        z_in1=w1+((inputvar[0]*v11)+(inputvar[1]*v12))
        z_in2=w2+((inputvar[0]*v21)+(inputvar[1]*v22))
        z_in3=w3+((inputvar[0]*v31)+(inputvar[1]*v32))
        z_in4=w4+((inputvar[0]*v41)+(inputvar[1]*v42))
        z_in5=w5+((inputvar[0]*v51)+(inputvar[1]*v52))
        
        #fungsi aktivasi
        z1=1/(1+pow(e,-(z_in1)))
        z2=1/(1+pow(e,-(z_in2)))
        z3=1/(1+pow(e,-(z_in3)))
        z4=1/(1+pow(e,-(z_in4)))
        z5=1/(1+pow(e,-(z_in5)))
        
        
        #output
        y_in=b0+((z1*bz1)+(z2*bz2)+(z3*bz3)+(z4*bz4)+(z5*bz5))
        result=1/(1+pow(e,-(y_in)))
        print("Hasilt persamaan: " + str(result))
        return result

    def RGBCount(self, filename, saveImage):
        global rgb1, rgb2, rgb3, rgb4, rgb5, rgb6, rgb7, rgb8, rgb9, rgb10
        img = cv2.imread(filename)
        rgb_read=cv2.imread(filename)
        height, width = img.shape[:2]
        #get rgb
        #point 1
        rgb1 = rgb_read[int(height * .5)-20,0+100]
        #point 2
        rgb2 = rgb_read[int(height * .5)-20,50+int((width-150) * .30)]
        #point 3
        rgb3 = rgb_read[int(height * .5)-20,50+int((width-150) * .55)]
        #point 4
        rgb4 = rgb_read[int(height * .5)-20,50+int((width-150) * .70)]
        #point 5
        rgb5 = rgb_read[int(height * .5)-20,50+int((width-150) * .80)]

        #point 6
        rgb6 = rgb_read[int(height * .5)+20,0+100]
        #point 7
        rgb7 = rgb_read[int(height * .5)+20,50+int((width-150) * .30)]
        #point 8
        rgb8 = rgb_read[int(height * .5)+20,50+int((width-150) * .55)]
        #point 9
        rgb9 = rgb_read[int(height * .5)+20,50+int((width-150) * .70)]
        #point 10
        rgb10 = rgb_read[int(height * .5)+20,50+int((width-150) * .80)]
        
        #point 1
        cv2.line(img, (0+100,int(height * .5)-20),(0+100,int(height * .5)-20),(0,0,255),5)
        #point 2
        cv2.line(img, (100+int((width-150) * .30),int(height * .5)-20),(100+int((width-150) * .30),int(height * .5)-20),(0,0,255),5)
        #point 3
        cv2.line(img, (100+int((width-150) * .55),int(height * .5)-20),(100+int((width-150) * .55),int(height * .5)-20),(0,0,255),5)
        #point 4
        cv2.line(img, (100+int((width-150) * .70),int(height * .5)-20),(100+int((width-150) * .70),int(height * .5)-20),(0,0,255),5)
        #point 5
        cv2.line(img, (100+int((width-150) * .80),int(height * .5)-20),(100+int((width-150) * .80),int(height * .5)-20),(0,0,255),5)

        #point 6
        cv2.line(img, (0+100,int(height * .5)+20),(0+100,int(height * .5)+20),(0,0,255),5)
        #point 7
        cv2.line(img, (100+int((width-150) * .30),int(height * .5)+20),(100+int((width-150) * .30),int(height * .5)+20),(0,0,255),5)
        #point 8
        cv2.line(img, (100+int((width-150) * .55),int(height * .5)+20),(100+int((width-150) * .55),int(height * .5)+20),(0,0,255),5)
        #point 9
        cv2.line(img, (100+int((width-150) * .70),int(height * .5)+20),(100+int((width-150) * .70),int(height * .5)+20),(0,0,255),5)
        #point 10
        cv2.line(img, (100+int((width-150) * .80),int(height * .5)+20),(100+int((width-150) * .80),int(height * .5)+20),(0,0,255),5)
        
        
        if saveImage==True:
            cv2.imwrite("data/rgbdaun/"+ str(self.lastNo)+".png", img)
        else:
            cv2.imwrite("rgbdaun.png", img)
        
        #image.set_from_file("rgbdaun.png")

        print("RGB Titik ke 1:" ,rgb1)
        print("RGB Titik ke 2:" ,rgb2)
        print("RGB Titik ke 3:" ,rgb3)
        print("RGB Titik ke 4:" ,rgb4)
        print("RGB Titik ke 5:" ,rgb5)
        print("RGB Titik ke 6:" ,rgb6)
        print("RGB Titik ke 7:" ,rgb7)
        print("RGB Titik ke 8:" ,rgb8)
        print("RGB Titik ke 9:" ,rgb9)
        print("RGB Titik ke 10:" ,rgb10)
        

        lblHasil.set_text("Titik ke 1: " + str(rgb1) + "\n" +
                          "Titik ke 2: " + str(rgb2) + "\n" +
                          "Titik ke 3: " + str(rgb3) + "\n" +
                          "Titik ke 4: " + str(rgb4) + "\n" +
                          "Titik ke 5: " + str(rgb5) + "\n" +
                          "Titik ke 6: " + str(rgb6) + "\n" +
                          "Titik ke 7: " + str(rgb7) + "\n" +
                          "Titik ke 8: " + str(rgb8) + "\n" +
                          "Titik ke 9: " + str(rgb9) + "\n" +
                          "Titik ke 10: " + str(rgb10))
        result = [rgb1, rgb2,rgb3,rgb4,rgb5, rgb6, rgb7, rgb8, rgb9, rgb10]
        return result

    def countKeliling(self, filename, saveImage):
        global kelilingAtas, kelilingBawah
        #reading the image 
        image = cv2.imread(filename)
        
        edged = cv2.Canny(image, 70, 150)
        
        count_keliling = cv2.countNonZero(edged)

        height, width = edged.shape[:2]
        # Let's get the starting pixel coordiantes (top left of cropped top)
        start_row, start_col = int(0), int(0)
        # Let's get the ending pixel coordinates (bottom right of cropped top)
        end_row, end_col = int(height * .5), int(width)
        cropped_top = edged[start_row:end_row , start_col:end_col]
        if saveImage==True:
            cv2.imwrite("data/kelilingdaun/"+str(self.lastNo)+"_atas.png",cropped_top)
        else:
            cv2.imwrite("countourTop.png",cropped_top)
        h_top, w_top = edged.shape[:2]
        cons_keliling = 33 #19cmx+(2*7cm) 7cm = 14cm/2 satu sisi tidak dihitung karna dibagi 2
        #dapatkan ukuan pixel per cm
        cons_divided = ((2*h_top) + w_top)/cons_keliling
        cropped_top = np.array(cropped_top)
        
        count_keliling = cv2.countNonZero(cropped_top)
        kelilingAtas=round(count_keliling/cons_divided)

        # Let's get the starting pixel coordiantes (top left of cropped bottom)
        start_row, start_col = int(height * .5), int(0)
        # Let's get the ending pixel coordinates (bottom right of cropped bottom)
        end_row, end_col = int(height), int(width)
        cropped_bot = edged[start_row:end_row , start_col:end_col]
        if saveImage==True:
            cv2.imwrite("data/kelilingdaun/"+str(self.lastNo)+"_bawah.png",cropped_bot)
        else:
            cv2.imwrite("countourBot.png",cropped_bot)
        h_bot, w_bot = cropped_bot.shape[:2]
        cropped_bot = np.array(cropped_bot)

        count_keliling = cv2.countNonZero(cropped_bot)
        kelilingBawah=round(count_keliling/cons_divided)

        print("Keliling daun bagian atas: ", kelilingAtas)
        print("Keliling daun bagian bawah: ", kelilingBawah)
        lblHasil.set_text("(" + str(kelilingAtas) + "," + str(kelilingBawah)+ ")")
        result = [kelilingAtas, kelilingBawah]
        return result
        

    def countAreaObject(self, filename, saveImage):
        global luasdauntop, luasdaunbot
        
        img = cv2.imread(filename)
        
        gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
        gray = cv2.medianBlur(gray, 5)
        #cv2.imwrite("gray.jpg",gray)
        ret, thresh = cv2.threshold(gray, 70, 255,cv2.THRESH_BINARY)
            
        height, width = thresh.shape[:2]
        # Let's get the starting pixel coordiantes (top left of cropped top)
        start_row, start_col = int(0), int(0)
        # Let's get the ending pixel coordinates (bottom right of cropped top)
        end_row, end_col = int(height * .5), int(width)
        cropped_top = thresh[start_row:end_row , start_col:end_col]

        h_top, w_top = cropped_top.shape[:2]
        
        cons_luas = 133 #19cmx7cm 7cm = 14cm/2
        #dapatkan ukuan pixel per cm
        cons_divided = (h_top * w_top)/cons_luas
        
        cropped_top = np.array(cropped_top)
        if saveImage==True:
            cv2.imwrite("data/luasdaun/"+str(self.lastNo)+"_atas.png", cropped_top)
        else:
            cv2.imwrite("cropped_top.png", cropped_top)

        count_white_top = cv2.countNonZero(cropped_top)
        luasdauntop = (h_top * w_top)- count_white_top
        luasdauntop = round(luasdauntop/cons_divided)      
        print("Luas daun bagian atas: ", luasdauntop)
        
        # Let's get the starting pixel coordiantes (top left of cropped bottom)
        start_row, start_col = int(height * .5), int(0)
        # Let's get the ending pixel coordinates (bottom right of cropped bottom)
        end_row, end_col = int(height), int(width)
        cropped_bot = thresh[start_row:end_row , start_col:end_col]
        h_bot, w_bot = cropped_bot.shape[:2]
        cropped_bot = np.array(cropped_bot)

        if saveImage==True:
            cv2.imwrite("data/luasdaun/"+str(self.lastNo)+"_bawah.png", cropped_bot)
        else:
            cv2.imwrite("cropped_bot.png", cropped_bot)
        count_white_bot = cv2.countNonZero(cropped_bot)
        luasdaunbot = (h_bot * w_bot)- count_white_bot
        luasdaunbot = round(luasdaunbot/cons_divided)
        
        
        print("Luas daun bagian bawah: ", luasdaunbot)
        lblHasil.set_text("(" + str(luasdauntop) + "," + str(luasdaunbot)+ ")")
        result = [luasdauntop,luasdaunbot]
        return result

    def thresholdingTest(self, filename):
        #reading the image 
        image = cv2.imread(filename)
        gray = cv2.cvtColor(image,cv2.COLOR_BGR2GRAY)
        gray = cv2.medianBlur(gray, 5)
        cv2.imwrite("gray.jpg",gray)
        for i in range(127):            
            ret, thresh = cv2.threshold(gray, i, 255,cv2.THRESH_BINARY)
            cv2.imwrite("thresh"+str(i)+".jpg",thresh)
            th2 = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY,11,2)
            cv2.imwrite("adaptivethresh"+str(i)+".jpg",thresh)
        
        print ("selesai")
    def background_remover4(self, filename):
        # Read image
        im_in = cv2.imread(filename, cv2.IMREAD_GRAYSCALE);
         
        # Threshold.
        # Set values equal to or above 220 to 0.
        # Set values below 220 to 255.
         
        th, im_th = cv2.threshold(im_in, 70, 255, cv2.THRESH_BINARY_INV);
         
        # Copy the thresholded image.
        im_floodfill = im_th.copy()
         
        # Mask used to flood filling.
        # Notice the size needs to be 2 pixels than the image.
        h, w = im_th.shape[:2]
        mask = np.zeros((h+2, w+2), np.uint8)
         
        # Floodfill from point (0, 0)
        cv2.floodFill(im_floodfill, mask, (0,0), 255);
         
        # Invert floodfilled image
        im_floodfill_inv = cv2.bitwise_not(im_floodfill)
        test = cv2.bitwise_and(im_floodfill)
        test2 = cv2.bitwise_and(im_floodfill_inv)
        # Combine the two images to get the foreground.
        im_out = im_th | im_floodfill_inv
         
        # Display images.
        cv2.imwrite("Thresholded Image.jpg", im_th)
        cv2.imwrite("Floodfilled Image.jpg", im_floodfill)
        cv2.imwrite("Inverted Floodfilled Image.jpg", im_floodfill_inv)
        cv2.imwrite("Foreground.jpg", im_out)
        cv2.imwrite("Foreground1.jpg", test)
        cv2.imwrite("Foreground2.jpg", test2)

    def background_remover3(self, img):
        stencil = np.zeros(img.shape).astype(img.dtype)
        #(_, cnts, _) = cv2.findContours(img.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        contours = [np.array([[100, 180], [200, 280], [200, 180]]), np.array([[280, 70], [12, 20], [80, 150]])]
        color = [255, 255, 255]
        cv2.fillPoly(stencil, contours, color)
        result = cv2.bitwise_and(img, stencil)
        cv2.imwrite("result.jpg", result)

    def background_remover2(self, filename):
        original_image = cv2.imread(filename)
        # 1 Convert to gray & Normalize
        gray_img = cv2.cvtColor(original_image, cv2.COLOR_BGR2GRAY)
        gray_img = sharp(cv2.get_contrasted(gray_img))
        gray_img = normalize(gray_img, None, 0, 255, cv2.NORM_MINMAX, cv2.CV_8UC1)
        cv2.imwrite("Gray.jpg", gray_img)

        # 2 Find Threshold
        gray_blur = cv2.GaussianBlur(gray_img, (7, 7), 0)
        adapt_thresh_im = cv2.adaptiveThreshold(gray_blur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 11, 1)
        max_thresh, thresh_im = cv2.threshold(gray_img, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)
        thresh = cv2.bitwise_or(adapt_thresh_im, thresh_im)

        # 3 Dilate
        gray = cv2.Canny(thresh, 88, 400, apertureSize=3)
        gray = cv2.dilate(gray, None, iterations=8)
        gray = cv2.erode(gray, None, iterations=8)
        cv2.imwrite("Trheshold.jpg", gray)

        # 4 Flood
        contours, _ = cv2.findContours(gray, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        contour_info = []
        for c in contours:
            contour_info.append((
                c,
                cv2.isContourConvex(c),
                cv2.contourArea(c),
            ))
        contour_info = sorted(contour_info, key=lambda c: c[2], reverse=True)
        max_contour = contour_info[0]
        holes = np.zeros(gray_img.shape, np.uint8)
        drawContours(holes, max_contour, 0, 255, -1)
        cv2.imwrite("Holes.jpg", holes)

        mask = cv2.GaussianBlur(holes, (15, 15), 0)
        mask = np.dstack([mask] * 3)  # Create 3-channel alpha mask

        mask = mask.astype('float32') / 255.0  # Use float matrices,
        img = original_image.astype('float32') / 255.0  # for easy blending
        masked = (mask * img) + ((1 - mask) * (0,0,1))  # Blend
        masked = (masked * 255).astype('uint8')

        cv2.imwrite("Maked.jpg", masked)

    def get_holes(self, image, thresh):
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

        im_bw = cv2.threshold(gray, thresh, 255, cv2.THRESH_BINARY)[1]
        im_bw_inv = cv2.bitwise_not(im_bw)

        (_, cnts, _) = cv2.findContours(im_bw_inv.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        for cnt in cnts:
            cv2.drawContours(im_bw_inv, [cnt], 0, 255, -1)
        cv2.imwrite("testss.jpg", im_bw_inv)
        nt = cv2.bitwise_not(im_bw)
        im_bw_inv = cv2.bitwise_or(im_bw_inv, nt)
        return im_bw_inv


    def remove_background(self, image, thresh, scale_factor=.25, kernel_range=range(1, 15), border=None):
        border = border or kernel_range[-1]

        holes = self.get_holes(image, thresh)
        small = cv2.resize(holes, None, fx=scale_factor, fy=scale_factor)
        bordered = cv2.copyMakeBorder(small, border, border, border, border, cv2.BORDER_CONSTANT)

        for i in kernel_range:
            kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (2*i+1, 2*i+1))
            bordered = cv2.morphologyEx(bordered, cv2.MORPH_CLOSE, kernel)

        unbordered = bordered[border: -border, border: -border]
        mask = cv2.resize(unbordered, (image.shape[1], image.shape[0]))
        fg = cv2.bitwise_and(image, image, mask=mask)
        return fg

    def background_remover(self, imgo):
        height, width = imgo.shape[:2]

        #Create a mask holder
        mask = np.zeros(imgo.shape[:2],np.uint8)

        #Grab Cut the object
        bgdModel = np.zeros((1,65),np.float64)
        fgdModel = np.zeros((1,65),np.float64)

        #Hard Coding the Rect… The object must lie within this rect.
        rect = (10,10,width-30,height-30)
        cv2.grabCut(imgo,mask,rect,bgdModel,fgdModel,5,cv2.GC_INIT_WITH_RECT)
        mask = np.where((mask==2)|(mask==0),0,1).astype('uint8')
        img1 = imgo*mask[:,:,np.newaxis]

        #Get the background
        background = imgo - img1

        #Change all pixels in the background that are not black to white
        background[np.where((background > [0,0,0]).all(axis = 2))] = [255,255,255]

        #Add the background and the image
        final = background + img1
        return final
        
    def getMultiImage(self):
        camera = cv2.VideoCapture(0)

        for i in range(10):
            print ("Taking image")
            retval, camera_capture = camera.read()
            file = "tmp.png"
            file = self.checkfile(file)
            cv2.imwrite(file, camera_capture)        
        del(camera)
        
    def checkfile(self, path):
         path      = os.path.expanduser(path)

         if not os.path.exists(path):
            return path

         root, ext = os.path.splitext(os.path.expanduser(path))
         dir       = os.path.dirname(root)
         fname     = os.path.basename(root)
         candidate = fname+ext
         index     = 0
         ls        = set(os.listdir(dir))
         while candidate in ls:
                 candidate = "{}_{}{}".format(fname,index,ext)
                 index    += 1
         return os.path.join(dir,candidate)


builder.connect_signals(mainWindowHandler())

Gtk.main()
